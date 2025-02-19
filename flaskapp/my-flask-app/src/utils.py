import pdfplumber
import pandas as pd
import re
import json
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from itertools import cycle

def process_pdf_file(filepath):
    """Process PDF file and return analysis results."""
    try:
        data = extract_data_from_pdf(filepath)
        df = pd.DataFrame(data)
        
        # Convert numerical columns
        numeric_columns = ["Enrollment Number", "Seat Number", "Total Marks", "Total"]
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Calculate percentage
        df['Percentage'] = df['Percentage'].fillna(0).round(2)
        
        # Get analysis results
        toppers_data = analyze_toppers(df)
        analysis_data = analyze_results(df)
        
        # Create Excel files
        excel_files = create_excel_files(data, toppers_data, analysis_data)
        
        return {
            'toppers': toppers_data,
            'analysis': analysis_data,
            'excel_files': excel_files
        }
        
    except Exception as e:
        print(f"Error processing PDF: {str(e)}")
        return None

def extract_data_from_pdf(pdf_path):
    """Extract structured data from a PDF result sheet."""
    data = []
    with pdfplumber.open(pdf_path) as pdf:
        current_semester = "Unknown"
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=2)
            if text:
                semester_match = re.search(r'RESULT SHEET FOR THE (.+? SEMESTER \(\w+\))', text, re.IGNORECASE)
                if semester_match:
                    current_semester = semester_match.group(1).strip()
                data.extend(parse_text(text, current_semester))
    return data

def parse_text(text, semester):
    """Parse extracted text from PDF."""
    records = []
    
    # Extract general information
    university = "Maharashtra State Board of Technical Education, Mumbai"
    institute = re.search(r'INSTITUTE : (.+?) COURSE', text)
    course = re.search(r'COURSE : (.+?)\n', text)
    exam_session = re.search(r'EXAMINATION HELD IN (.+?) \(', text)
    result_date_match = re.search(r'Result Date : (\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    total_match = re.search(r'Total Marks : (\d+)', text)
    
    institute = institute.group(1).strip() if institute else "Unknown"
    course = course.group(1).strip() if course else "Unknown"
    exam_session = exam_session.group(1).strip() if exam_session else "Unknown"
    result_date = result_date_match.group(1) if result_date_match else "Unknown"
    total_possible = int(total_match.group(1)) if total_match else 1000
    
    # Use appropriate pattern based on semester type
    if "(K)" in semester:
        student_pattern = re.compile(
            r'(?P<seat_no>\d{6})\s+'
            r'(?P<enroll_no>\d{10})\s+'
            r'(?P<name>[A-Za-z ]{5,})\s*'
            r'(?P<app_code>[A-Z ]{2,15})?\s*'
            r'Total\s*:\s*(?P<total_marks>\d+)',
            re.DOTALL
        )
    else:
        student_pattern = re.compile(
            r'(?P<seat_no>\d{6})\s+(?P<enroll_no>\d{10})\s+'
            r'(?P<name>(?:[A-Z]+(?:\s+[A-Z]+)*))\s+'
            r'(?P<app_code>[A-Z]+(?:\s+[A-Z0-9]+)*)?\s+'
            r'.*?Total\s*:\s*(?P<total_marks>\d+)',
            re.DOTALL
        )
    
    for match in student_pattern.finditer(text):
        seat_no = match.group("seat_no")
        enroll_no = match.group("enroll_no")
        name = match.group("name").strip()
        total_marks = int(match.group("total_marks"))
        
        # Calculate percentage
        percentage = (total_marks / total_possible) * 100 if total_possible else 0
        
        result_status_match = re.search(rf'{seat_no}.*?Result\s*:\s*([A-Z .]+)', text, re.DOTALL)
        result_status = result_status_match.group(1).strip() if result_status_match else "UNKNOWN"
        
        record = {
            "University Name": university,
            "Institute Name": institute,
            "Course Name": course,
            "Semester": semester,
            "Exam Session": exam_session,
            "Result Date": result_date,
            "Student Name": name,
            "Enrollment Number": int(enroll_no),
            "Seat Number": int(seat_no),
            "Total Marks": total_marks,
            "Total": total_possible,
            "Percentage": round(percentage, 2),
            "Result Status": result_status
        }
        records.append(record)
    
    return records

def analyze_toppers(df):
    """Identify toppers by total marks and average percentage per course per year."""
    # Add academic year
    academic_year = "2023-24"
    
    # Extract semester numbers
    df["Semester Number"] = df["Semester"].apply(lambda x: {
        "FIRST": 1, "SECOND": 2, "THIRD": 3,
        "FOURTH": 4, "FIFTH": 5, "SIXTH": 6
    }.get(next((k for k in ["FIRST", "SECOND", "THIRD", "FOURTH", "FIFTH", "SIXTH"] 
                if k in x.upper()), None), None))
    
    # Determine year
    df["Year"] = df["Semester Number"].apply(
        lambda x: "First Year" if x in [1, 2] else 
                 ("Second Year" if x in [3, 4] else "Third Year"))
    
    # Calculate yearly totals and averages
    df_yearly = df.groupby(["Course Name", "Student Name", "Year"]).agg({
        "Total Marks": "sum",  # Sum of both semester marks
        "Total": "sum",        # Sum of total possible marks
        "Percentage": "mean"   # Average percentage across semesters
    }).reset_index()
    
    # Round percentage to 2 decimal places
    df_yearly["Percentage"] = df_yearly["Percentage"].round(2)
    
    # Get top 3 students per course and year based on percentage
    toppers = df_yearly.sort_values(
        ["Course Name", "Year", "Percentage"], 
        ascending=[True, True, False]
    ).groupby(["Course Name", "Year"]).head(3)
    
    # Add academic year and rank columns
    final_toppers = []
    for (course, year), group in toppers.groupby(["Course Name", "Year"]):
        for rank, (_, row) in enumerate(group.iterrows(), 1):
            final_toppers.append({
                "Academic Year": academic_year,
                "Course Name": row["Course Name"],
                "Student Name": row["Student Name"],
                "Year": row["Year"],
                "Rank": f"Rank {rank}",
                "Total Marks": row["Total Marks"],
                "Percentage": row["Percentage"]
            })
    
    return final_toppers

def analyze_results(df):
    """Perform semester-wise result analysis per course."""
    result_analysis = {}
    for (course, semester), group in df.groupby(["Course Name", "Semester"]):
        total_students = len(group)
        # Updated result status patterns
        distinction = len(group[group["Result Status"].str.contains("FIRST CLASS DIST.", na=False)])
        first_class = len(group[group["Result Status"].str.contains("FIRST CLASS TCALSE|FIRST CLASS CON TCALSE", na=False)])
        second_class = len(group[group["Result Status"].str.contains("SECOND CLASS", na=False)])
        fail = len(group[group["Result Status"].str.contains("FAIL", na=False)])
        atkt = len(group[group["Result Status"].str.contains("A.T.K.T.|F.T.", na=False)])
        passed = total_students - fail
        first_class_or_more = distinction + first_class
        
        # Maintain the same structure for JSON output
        result_analysis.setdefault(course, {})[semester] = {
            "Total Students": total_students,
            "Distinction": distinction,
            "First Class": first_class,
            "First Class or More": first_class_or_more,
            "Second Class": second_class,
            "Fail": fail,
            "ATKT": atkt,
            "Pass": passed,
            "Pass Percentage": round((passed / total_students) * 100, 2) if total_students > 0 else 0,
            "Grade Distribution": {
                "Distinction": round((distinction / total_students) * 100, 2) if total_students > 0 else 0,
                "First Class": round((first_class / total_students) * 100, 2) if total_students > 0 else 0,
                "Second Class": round((second_class / total_students) * 100, 2) if total_students > 0 else 0,
                "ATKT": round((atkt / total_students) * 100, 2) if total_students > 0 else 0,
                "Fail": round((fail / total_students) * 100, 2) if total_students > 0 else 0
            }
        }
    print("data send from result analysis")
    return result_analysis

def create_excel_files(raw_data, toppers_data, analysis_data):
    """Create Excel files with enhanced formatting and structure."""
    try:
        excel_files = {}

        
        # Define border style at the start
        border_style = Side(style='thin', color='000000')
        
        # Update the color palette with light/medium variants for each course
        course_colors = {
            'header': '1E88E5',      # Blue header
            'pass': '4CAF50',        # Green
            'fail': 'F44336',        # Red
            'course_palette': {
                'Computer Engineering': {'light': 'FFE0B2', 'medium': 'FFB74D'},      # Light Orange/Peach
                'Information Technology': {'light': 'E1BEE7', 'medium': 'CE93D8'},    # Light Purple/Lavender
                'Civil Engineering': {'light': 'B3E5FC', 'medium': '81D4FA'},         # Light Blue/Sky
                'Mechanical Engineering': {'light': 'C8E6C9', 'medium': 'A5D6A7'},    # Light Green/Mint
                'Electrical Engineering': {'light': 'FFCCBC', 'medium': 'FFAB91'},    # Light Coral/Salmon
                'Electronics Engineering': {'light': 'D7CCC8', 'medium': 'BCAAA4'},   # Light Brown/Taupe
                'Default': {'light': 'F5F5F5', 'medium': 'E0E0E0'}                   # Light Grey
            }
        }

        def apply_formatting(worksheet, course=None):
            """Apply formatting to worksheet with alternating light/medium colors"""
            if not worksheet:
                return
                
            # Format headers
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color=course_colors['header'], 
                                      end_color=course_colors['header'], 
                                      fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=border_style, bottom=border_style, 
                                   left=border_style, right=border_style)
            
            # Group rows by course for alternating colors
            current_course = None
            course_row_count = 0
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                row_course = row[0].value if course is None else course
                
                # Reset counter when course changes
                if row_course != current_course:
                    current_course = row_course
                    course_row_count = 0
                
                # Get course colors (default if course not in palette)
                course_colors_dict = course_colors['course_palette'].get(
                    row_course, 
                    course_colors['course_palette']['Default']
                )
                
                # Alternate between light and medium
                color = course_colors_dict['medium'] if course_row_count % 2 == 0 else course_colors_dict['light']
                
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(top=border_style, bottom=border_style, 
                                          left=border_style, right=border_style)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        # Use white text for dark backgrounds
                        cell.font = Font(color='FFFFFF' if course_row_count % 2 == 0 else '000000')
                
                course_row_count += 1

            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Create Excel files with error handling
        print("exel file one")
        try:
            # 1. Raw Data Excel
            
            raw_excel = BytesIO()
            with pd.ExcelWriter(raw_excel, engine='openpyxl') as writer:
                pd.DataFrame(raw_data).to_excel(writer, sheet_name='Extracted Data', index=False)
                apply_formatting(writer.sheets['Extracted Data'])
            raw_excel.seek(0)
            excel_files['raw_data'] = raw_excel

            # 2. Toppers Excel
            print("exel file 2one")
            toppers_excel = BytesIO()
            
            with pd.ExcelWriter(toppers_excel, engine='openpyxl') as writer:
                pd.DataFrame(toppers_data).to_excel(writer, sheet_name='Toppers', index=False)
                worksheet = writer.sheets['Toppers']
                apply_formatting(worksheet)
                
                # Highlight top performers
                for row in worksheet.iter_rows(min_row=2):
                    if row[-1].value:  # Check if percentage exists
                        percentage = float(str(row[-1].value).replace('%', ''))
                        if percentage >= 75:
                            for cell in row:
                                cell.font = Font(color=course_colors['pass'], bold=True)
            toppers_excel.seek(0)
            excel_files['toppers'] = toppers_excel

            # 3. Analysis Excel
            print("exel file 3one")
            analysis_excel = BytesIO()
            with pd.ExcelWriter(analysis_excel, engine='openpyxl') as writer:
                # Create and format analysis sheet
                analysis_rows = [
                    {
                        'Course': course,
                        'Semester': semester,
                        'Total Students': stats['Total Students'],
                        'Pass': stats['Pass'],
                        'Pass Percentage': f"{stats['Pass Percentage']}%",
                        'Distinction': stats['Distinction'],
                        'First Class': stats['First Class'],
                        'Second Class': stats['Second Class'],
                        'ATKT': stats['ATKT'],
                        'Fail': stats['Fail']
                    }
                    for course, semesters in analysis_data.items()
                    for semester, stats in semesters.items()
                ]
                
                pd.DataFrame(analysis_rows).to_excel(writer, sheet_name='Analysis', index=False)
                apply_formatting(writer.sheets['Analysis'])
                
                # Format grade distribution sheet
                dist_rows = [
                    {
                        'Course': course,
                        'Semester': semester,
                        **{k: f"{v}%" for k, v in stats['Grade Distribution'].items()}
                    }
                    for course, semesters in analysis_data.items()
                    for semester, stats in semesters.items()
                ]
                
                pd.DataFrame(dist_rows).to_excel(writer, sheet_name='Grade Distribution', index=False)
                apply_formatting(writer.sheets['Grade Distribution'])
            
            analysis_excel.seek(0)
            excel_files['analysis'] = analysis_excel

            return excel_files

        except Exception as e:
            print(f"Error creating Excel files: {str(e)}")
            return None

    except Exception as e:
        print(f"Error in create_excel_files: {str(e)}")
        return None